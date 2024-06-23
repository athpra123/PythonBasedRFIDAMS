from ast import Return
from cProfile import label
from re import A
import tkinter as tk
from tkinter import messagebox
import sqlite3 as sql
from tkinter import *
from functools import partial
from turtle import bgcolor
from typing_extensions import Self
import openpyxl 
from PIL import Image, ImageTk
import serial
import pyttsx3
import time

class AttendanceManager(tk.Tk):
	def __init__(self,*args,**kwargs):
		tk.Tk. __init__(self,*args,**kwargs)
		container=tk.Frame(self)
		
		container.pack(side="top",fill="both",expand=True)
		container.grid_rowconfigure(0,weight=1)
		container.grid_columnconfigure(0,weight=1)
		
		self.frames=dict() 	
		for F in (ShowStatus,Student_teacher,TakeattendanceEdc,TakeattendanceLnt,TakeattendanceSns,TakeattendanceMal,TakeattendanceOop,CourseloginEdc,CourseloginOop,CourseloginSns,CourseloginMal,CourseloginLnt,Courselist,FirstPage,RegisterPage,LoginPage,StartPage,NewRecord,ManageAttendance,DeleteRecord,AddSubjects):
			frame=F(container,self)
			self.frames[F]=frame
			frame.grid(row=0,column=0,sticky="nsew")
		
		self.show_frame(Student_teacher)
		
	def show_frame(self,cont):
		frame=self.frames[cont]
		frame.tkraise()

class Student_teacher(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Welcome to the Attendance Manager Portal \n\n\n Who is the User?",font=("Caslon",16),background="#ffffe4")
		label1.pack(padx=10,pady=25)
		bu1=tk.Button(self,text="Teacher",font=("Caslon",12),height=1,width=16,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:[controller.show_frame(Courselist),self.speaker('Select The Subject You Teach To Proceed.')])
		bu2=tk.Button(self,text="Student",font=("Caslon",12),height=1,width=16,default='active',bg="#2983cc",fg='white', cursor='boat',command=lambda:[controller.show_frame(FirstPage),self.speaker('Register Or Login to continue.')])
		bu1.pack(padx=10,pady=25)
		bu2.pack(padx=10,pady=25)
	def speaker(self,text):
		pyttsx3.speak(text)
		

class Courselist(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		bt_sns = tk.Button(self,text="Signals And Systems",font=("Caslon",12),height=1,width=25,bg="#2983cc",fg='white',default='active', cursor='boat' ,command=lambda:controller.show_frame(CourseloginSns))
		bt_edc = tk.Button(self,text="Electronic Devices",font=("Caslon",12),height=1,width=25,bg="#2983cc",fg='white',default='active', cursor='boat' ,command=lambda:controller.show_frame(CourseloginEdc))
		bt_mal = tk.Button(self,text="Linear Algebra",font=("Caslon",12),height=1,width=25,bg="#2983cc",fg='white',default='active', cursor='boat' ,command=lambda:controller.show_frame(CourseloginMal))
		bt_oop = tk.Button(self,text="Object Oriented Programming",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:controller.show_frame(CourseloginOop))
		bt_lnt = tk.Button(self,text="Linear Network Theory",font=("Caslon",12),height=1,width=25,bg="#2983cc",fg='white',default='active', cursor='boat' ,command=lambda:controller.show_frame(CourseloginLnt))
		bt_back=tk.Button(self,text="Back",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat',command=lambda:controller.show_frame(Student_teacher))
		bt_quit=tk.Button(self,text="Quit",font=("Caslon",12),height=1,width=16,bg="red",fg='white',default='active', cursor='boat',command=quit)
		bt_sns.pack(padx=10,pady=25)
		bt_edc.pack(padx=10,pady=25)
		bt_mal.pack(padx=10,pady=25)
		bt_oop.pack(padx=10,pady=25)
		bt_lnt.pack(padx=10,pady=25)
		bt_back.pack(padx=10,pady=25)
		bt_quit.pack(padx=10,pady=25)

class CourseloginEdc(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		
		label1=tk.Label(self,text="Enter your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(Courselist))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
	
		bt2=tk.Button(self,text="Login",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(TakeattendanceEdc))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)

class CourseloginSns(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		
		label1=tk.Label(self,text="Enter your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(Courselist))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
	
		bt2=tk.Button(self,text="Login",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(TakeattendanceSns))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)

class CourseloginMal(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		
		label1=tk.Label(self,text="Enter your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(Courselist))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
	
		bt2=tk.Button(self,text="Login",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(TakeattendanceMal))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)

class CourseloginOop(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		
		label1=tk.Label(self,text="Enter your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(Courselist))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
	
		bt2=tk.Button(self,text="Login",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(TakeattendanceOop))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)

class CourseloginLnt(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		
		label1=tk.Label(self,text="Enter your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(Courselist))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
	
		bt2=tk.Button(self,text="Login",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(TakeattendanceLnt))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)

class TakeattendanceEdc(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		butt1 = tk.Button(self,text="Start Taking Attendence ",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:self.Running())
		butt1.pack(padx=10,pady=25)
		butt2 = tk.Button(self,text="Go Back",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:controller.show_frame(Courselist))
		butt2.pack(padx=10,pady=25)

	def Running(self):
		from serial.tools import list_ports
		def rfid():

			ser = serial.Serial()
			ser.baudrate = 9600

			try:
				ser.port = 'COM4'
			except:
				ser.port = 'COM5'
			ser.open()
			RFID_Data=ser.readline() 
			if RFID_Data:
				RFID_Data = RFID_Data.decode()  #Decode arduino Serial
				RFID_Data = RFID_Data.strip()   #Strip Arduino Data to remove string
				RFID_Data=int(RFID_Data);       #Convert the Data to Int
				return(RFID_Data)
		t = 0
		x1 = time.time()
		Go = True
		while Go:

			data = rfid()
			print(data)
			path = r"C:\Users\kartik\Desktop\GUI stuff\EDC.xlsx"
			wb_obj = openpyxl.load_workbook(path)
			sheet = wb_obj.active   
		
			for i in range(1,sheet.max_row+1):

					cell_obj = sheet.cell(row = i, column = 3).value
					if(cell_obj == None):
						continue
					else:
							lst = list(cell_obj)
							lst.pop()
							cell_obj1 = "".join(lst)
							if(data == int(cell_obj1)):
								sheet.cell(row = i, column = sheet.max_column + 1 - t).value = 'Present'
								print('Value Assigned!')
								self.speaker('Attendance recorded!')
								t = t + 1
								break
			

			wb_obj.save(path)
			x2 = time.time()
			if((x2-x1)>15):
				break

	def speaker(self,text):
		pyttsx3.speak(text)

class TakeattendanceSns(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		butt1 = tk.Button(self,text="Start Taking Attendence ",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:self.Running())
		butt1.pack(padx=10,pady=25)
		butt2 = tk.Button(self,text="Go Back",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:controller.show_frame(Courselist))
		butt2.pack(padx=10,pady=25)

	def Running(self):
		from serial.tools import list_ports
		def rfid():

			ser = serial.Serial()
			ser.baudrate = 9600

			try:
				ser.port = 'COM4'
			except:
				ser.port = 'COM5'
			ser.open()
			RFID_Data=ser.readline() 
			if RFID_Data:
				RFID_Data = RFID_Data.decode()  #Decode arduino Serial
				RFID_Data = RFID_Data.strip()   #Strip Arduino Data to remove string
				RFID_Data=int(RFID_Data);       #Convert the Data to Int
				return(RFID_Data)
		t = 0
		x1 = time.time()
		Go =True
		while Go:

			data = rfid()
			print(data)
			path = r"C:\Users\kartik\Desktop\GUI stuff\SNS.xlsx"
			wb_obj = openpyxl.load_workbook(path)
			sheet = wb_obj.active   
		
			for i in range(1,sheet.max_row+1):

					cell_obj = sheet.cell(row = i, column = 3).value
					if(cell_obj == None):
						continue
					else:
							lst = list(cell_obj)
							lst.pop()
							cell_obj1 = "".join(lst)
							if(data == int(cell_obj1)):
								sheet.cell(row = i, column = sheet.max_column + 1 - t).value = 'Present'
								print('Value Assigned!')
								self.speaker('Attendance recorded!')
								t = t + 1
								break

			wb_obj.save(path)
			x2 = time.time()
			if((x2-x1)>15):
				break
	def speaker(self,text):
		pyttsx3.speak(text)

class TakeattendanceLnt(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		butt1 = tk.Button(self,text="Start Taking Attendence ",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:self.Running())
		butt1.pack(padx=10,pady=25)
		butt2 = tk.Button(self,text="Go Back",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:controller.show_frame(Courselist))
		butt2.pack(padx=10,pady=25)

	def Running(self):
		from serial.tools import list_ports
		def rfid():

			ser = serial.Serial()
			ser.baudrate = 9600

			try:
				ser.port = 'COM4'
			except:
				ser.port = 'COM5'
			ser.open()
			RFID_Data=ser.readline() 
			if RFID_Data:
				RFID_Data = RFID_Data.decode()  #Decode arduino Serial
				RFID_Data = RFID_Data.strip()   #Strip Arduino Data to remove string
				RFID_Data=int(RFID_Data);       #Convert the Data to Int
				return(RFID_Data)
		t = 0
		x1 = time.time()
		Go = True
		while Go:

			data = rfid()
			print(data)
			path = r"C:\Users\kartik\Desktop\GUI stuff\LNT.xlsx"
			wb_obj = openpyxl.load_workbook(path)
			sheet = wb_obj.active		
			for i in range(1,sheet.max_row+1):

					cell_obj = sheet.cell(row = i, column = 3).value
					if(cell_obj == None):
						continue
					else:
							lst = list(cell_obj)
							lst.pop()
							cell_obj1 = "".join(lst)
							if(data == int(cell_obj1)):
								sheet.cell(row = i, column = sheet.max_column + 1 - t).value = 'Present'
								print('Value Assigned!')
								self.speaker('Attendance recorded!')
								t = t + 1
								break

			wb_obj.save(path)
			x2 = time.time()
			if((x2-x1)>15):
				break
			
	def speaker(self,text):
		pyttsx3.speak(text)

class TakeattendanceOop(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		butt1 = tk.Button(self,text="Start Taking Attendence ",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:self.Running())
		butt1.pack(padx=10,pady=25)
		butt2 = tk.Button(self,text="Go Back",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:controller.show_frame(Courselist))
		butt2.pack(padx=10,pady=25)

	def Running(self):
		from serial.tools import list_ports
		def rfid():

			ser = serial.Serial()
			ser.baudrate = 9600

			try:
				ser.port = 'COM4'
			except:
				ser.port = 'COM5'
			ser.open()
			RFID_Data=ser.readline() 
			if RFID_Data:
				RFID_Data = RFID_Data.decode()  #Decode arduino Serial
				RFID_Data = RFID_Data.strip()   #Strip Arduino Data to remove string
				RFID_Data=int(RFID_Data);       #Convert the Data to Int
				return(RFID_Data)
		t = 0
		x1 = time.time()
		Go = True
		while Go:

			data = rfid()
			print(data)
			path = r"C:\Users\kartik\Desktop\GUI stuff\LNT.xlsx"
			wb_obj = openpyxl.load_workbook(path)
			sheet = wb_obj.active   
			for i in range(1,sheet.max_row+1):

					cell_obj = sheet.cell(row = i, column = 3).value
					if(cell_obj == None):
						continue
					else:
							lst = list(cell_obj)
							lst.pop()
							cell_obj1 = "".join(lst)
							#print(data == cell_obj1)
							if(data == int(cell_obj1)):
								sheet.cell(row = i, column = sheet.max_column + 1 - t).value = 'Present'
								print('Value Assigned!')
								self.speaker('Attendance recorded!')
								t = t + 1
								break
			x2 = time.time()
			if((x2-x1)>15):
				break
			wb_obj.save(path)

	def speaker(self,text):
		pyttsx3.speak(text)

class TakeattendanceMal(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		butt1 = tk.Button(self,text="Start Taking Attendence ",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:self.Running())
		butt1.pack(padx=10,pady=25)
		butt2 = tk.Button(self,text="Go Back",font=("Caslon",12),height=1,width=25,default='active',bg="#2983cc",fg='white', cursor='boat' ,command=lambda:controller.show_frame(Courselist))
		butt2.pack(padx=10,pady=25)

	def Running(self):
		from serial.tools import list_ports
		def rfid():

			ser = serial.Serial()
			ser.baudrate = 9600

			try:
				ser.port = 'COM4'
			except:
				ser.port = 'COM5'

			ser.open()
			RFID_Data=ser.readline() 
			if RFID_Data:
				RFID_Data = RFID_Data.decode()  #Decode arduino Serial
				RFID_Data = RFID_Data.strip()   #Strip Arduino Data to remove string
				RFID_Data=int(RFID_Data);       #Convert the Data to Int
				return(RFID_Data)
		t = 0
		Go = True
		x1 = time.time()
		while Go:

			data = rfid()
			print(data)
			path = r"C:\Users\kartik\Desktop\GUI stuff\MATHS.xlsx"
			wb_obj = openpyxl.load_workbook(path)
			sheet = wb_obj.active  
			#sheet.cell(row = 1, column = sheet.max_column + 1).value =  time.strftime("%d/%m/%Y") 
			for i in range(1,sheet.max_row + 1):

					cell_obj = sheet.cell(row = i, column = 3).value
					if(cell_obj == None):
						continue
					else:
							lst = list(cell_obj)
							lst.pop()
							cell_obj1 = "".join(lst)
							#print(data == cell_obj1)
							if(data == int(cell_obj1)):
								sheet.cell(row = i, column = sheet.max_column + 1 - t).value = 'Present'
								print('Value Assigned!')
								self.speaker('Attendance recorded!')
								t = t + 1
								break

			wb_obj.save(path)
			x2 = time.time()
			if((x2-x1)>15):
				break
	def speaker(self,text):
		pyttsx3.speak(text)

class FirstPage(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Register or Login to continue",font=("Caslon",16),background="#ffffe4")
		label1.pack(padx=10,pady=25)
		bt1=tk.Button(self,text="Register",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat' ,command=lambda:controller.show_frame(RegisterPage))
		bt2=tk.Button(self,text="Login",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat',command=lambda:controller.show_frame(LoginPage))
		bt3=tk.Button(self,text="Back",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat',command=lambda:controller.show_frame(Student_teacher))
		bt4=tk.Button(self,text="Quit",font=("Caslon",12),height=1,width=16,bg="red",fg='white',default='active', cursor='boat',command=quit)
		bt1.pack(padx=10,pady=25)
		bt2.pack(padx=10,pady=25)
		bt3.pack(padx=10,pady=25)
		bt4.pack(padx=10,pady=25)
		

class RegisterPage(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Choose your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(FirstPage))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
		bt2=tk.Button(self,text="Register",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(StartPage))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)

class LoginPage(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Enter your Username and Password: ",font=("Caslon",16),background="#ffffe4")
		bt1=tk.Button(self,text="Back to options",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",10),height=0,width=12,default='active',command=lambda:controller.show_frame(FirstPage))
		label1.pack(padx=10,pady=25)
		username = StringVar()
		lb2=tk.Label(self,text="Username: ",font=("Caslon",12),background="#ffffe4")
		txt1=tk.Entry(self, textvariable=username)
		lb2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		password = StringVar()
		lb3=tk.Label(self, text="Password: ",font=("Caslon",12),background="#ffffe4")
		txt2=tk.Entry(self, textvariable=password, show='*')
		lb3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
		bt2=tk.Button(self,text="Login",bg="lightblue",font=("Caslon",10),height=0,width=12,default='active',command=lambda: controller.show_frame(StartPage))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)
		
class StartPage(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent) 
		self.configure(background="#ffffe4")
		bt1=tk.Button(self,text="Add new record",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat' ,command=lambda:controller.show_frame(NewRecord))
		bt2=tk.Button(self,text="Manage attendance",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat',command=lambda:controller.show_frame(ManageAttendance))
		bt3=tk.Button(self,text="Delete record",font=("Caslon",12),height=1,width=16,bg="#2983cc",fg='white',default='active', cursor='boat',command=lambda:controller.show_frame(DeleteRecord))
		bt5=tk.Button(self,text="Quit",font=("Caslon",12),height=1,width=16,bg="red",fg='white',default='active', cursor='boat',command=quit)
		bt1.pack(padx=10,pady=25)
		bt2.pack(padx=10,pady=25)
		bt3.pack(padx=10,pady=25)
		bt5.pack(padx=10,pady=25)

class NewRecord(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="New Record",font=("Caslon",12),background="#ffffe4")
		label2=tk.Label(self,text="Old Record will be deleted, continue?",font=("Caslon",12),background="#ffffe4")
		
		bt2=tk.Button(self,text="Yes",font=("Caslon",12),bg="#2983cc",fg='white', cursor='boat',height=1,width=17,default='active',command=lambda:[controller.show_frame(AddSubjects),self.speaker('Enter the required fields for entry of data')])
		bt3=tk.Button(self,text="No",font=("Caslon",12),bg="#2983cc",fg='white', cursor='boat',height=1,width=17,default='active',command=lambda:controller.show_frame(StartPage))
		label1.pack(padx=10,pady=25)
		label2.pack(padx=10,pady=25)
		bt2.pack(padx=10,pady=25)
		bt3.pack(padx=10,pady=25)
	def speaker(self,text):
		pyttsx3.speak(text)
		

class ManageAttendance(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Manage Attendance",font=("Caslon",16),background="#ffffe4")
		label1.pack(padx=10,pady=25)
		bt2=tk.Button(self,text="Show status of Attendance", bg="#2983cc",fg='white', cursor='boat',font=("Caslon",12),height=1,width=20,default='active',command=lambda:[controller.show_frame(ShowStatus),self.speaker('Enter the roll number and select the subject to continue!')])	
		bt1=tk.Button(self,text="Back to home",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",12),height=1,width=20,default='active',command=lambda:controller.show_frame(StartPage))
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)
	def speaker(self,text):
		pyttsx3.speak(text)

class ShowStatus(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Enter Roll Number ",font=("Caslon",16),background="#ffffe4")
		txt1 = tk.Text(self,font=("Caslon",16),width=48,height=2)

		but1=tk.Button(self,text="Signals and Systems",bg="lightblue",font=("Caslon",12),height=1,width=25,default='active',command=lambda:self.show_sns(txt1.get("1.0",tk.END),controller))
		but2=tk.Button(self,text="Electronic Devices",bg="lightblue",font=("Caslon",12),height=1,width=25,default='active',command=lambda:self.show_edc(txt1.get("1.0",tk.END),controller))
		but3=tk.Button(self,text="Linear Network Theory",bg="lightblue",font=("Caslon",12),height=1,width=25,default='active',command=lambda:self.show_lnt(txt1.get("1.0",tk.END),controller))
		but4=tk.Button(self,text="Mathematics",bg="lightblue",font=("Caslon",12),height=1,width=25,default='active',command=lambda:self.show_mat(txt1.get("1.0",tk.END),controller))
		but5=tk.Button(self,text="Object Oriented Programming",bg="lightblue",font=("Caslon",12),height=1,width=25,default='active',command=lambda:self.show_oop(txt1.get("1.0",tk.END),controller))
		but6=tk.Button(self,text="Back to Home",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",12),height=1,width=25,default='active',command=lambda:controller.show_frame(ManageAttendance))
		label1.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		but1.pack(padx=10,pady=25)
		but2.pack(padx=10,pady=25)
		but3.pack(padx=10,pady=25)
		but4.pack(padx=10,pady=25)
		but5.pack(padx=10,pady=25)
		but6.pack(padx=10,pady=25)
		
	
	def show_sns(self,t,controller):
		path1 = r"C:\Users\kartik\Desktop\GUI stuff\SNS.xlsx"
		wb_obj1 = openpyxl.load_workbook(path1) 
		sheet1 = wb_obj1.active
		cnt1 = 0
		for i in range(1,200):
			cell_obj_name = sheet1.cell(row = i, column = 4 ).value
			

			if(cell_obj_name == None):
				continue
			else:
				if(cell_obj_name == None):
					for j in range(6, sheet1.max_column + 1):
						if(sheet1.cell(row = i, column = j).value != None):
							cnt1 = cnt1 + 1
		messagebox.showinfo("Attendance Report!","Number of classes attended is  " + str(cnt1+1))
		wb_obj1.save(path1)

	def show_oop(self,t,controller):
		path2 = r"D:\Software\Visual Studio\Visual Studio Programs\Excel\OOPS.xlsx"
		wb_obj2 = openpyxl.load_workbook(path2)
		sheet2 = wb_obj2.active 
		cnt1 = 0
		for i in range(1,200):
			cell_obj_name = sheet2.cell(row = i, column = 4 ).value
			if(cell_obj_name == None):
				continue
			else:
				if(cell_obj_name == None):
					for j in range(6,sheet2.max_column):
						if(sheet2.cell(row = i, column = j) != None):
							cnt1 = cnt1 + 1
		messagebox.showinfo("Attendance Report!","Number of classes attended is" + str(cnt1))
		wb_obj2.save(path2)
	
	def show_edc(self,t,controller):
		path3 = r"C:\Users\kartik\Desktop\GUI stuff\EDC.xlsx"
		wb_obj3 = openpyxl.load_workbook(path3) 
		sheet3 = wb_obj3.active
		cnt1 = 0
		for i in range(1,200):
			cell_obj_name = sheet3.cell(row = i, column = 4 ).value
			if(cell_obj_name == None):
				continue
			else:
				if(cell_obj_name == t):
					for j in range(6,sheet3.max_column+1):
						if(sheet3.cell(row = i, column = j).value != None):
							cnt1 = cnt1 + 1
			

		messagebox.showinfo("Attendance Report!","Number of classes attended is  " + str(cnt1))
		wb_obj3.save(path3)
	
	def show_lnt(self,t,controller):
		path4 = r"C:\Users\kartik\Desktop\GUI stuff\LNT.xlsx"
		wb_obj4 = openpyxl.load_workbook(path4) 
		sheet4 = wb_obj4.active
		cnt1 = 0
		for i in range(1,200):

			cell_obj_name = sheet4.cell(row = i, column = 4 ).value

			if(cell_obj_name == None):
				continue
			else:
				if(cell_obj_name == t):
					for j in range(6,sheet4.max_column+1):
							if(sheet4.cell(row = i, column = j).value != None):
								cnt1 = cnt1 + 1
							
			
		messagebox.showinfo("Attendance Report!","Number of classes attended is  " + str(cnt1))
		wb_obj4.save(path4)

	def show_mat(self,t,controller):
		path5 = r"D:\Software\Visual Studio\Visual Studio Programs\Excel\MATHS.xlsx"
		wb_obj5 = openpyxl.load_workbook(path5) 
		sheet5 = wb_obj5.active
		cnt1 = 0
		for i in range(1,200):
			cell_obj_name = sheet5.cell(row = i, column = 4 ).value
			if(cell_obj_name == None):
				continue
			else:
				if(cell_obj_name == t):
					for j in range(6,sheet5.max_column):
						if(sheet5.cell(row = i, column = j).value != None):
							cnt1 = cnt1 + 1

		messagebox.showinfo("Attendance Report!","Number of classes attended is" + str(cnt1))
		wb_obj5.save(path5)

class DeleteRecord(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Delete Record",font=("Caslon",16),background="#ffffe4")
		label2=tk.Label(self,text="Enter the Name of the student to delete his record! ",font=("Caslon",12),background="#ffffe4")
		txt1 = tk.Text(self,font=("Caslon",16),width=48,height=2) #name
		label3=tk.Label(self,text="Enter the Roll Number of the student to delete his record! ",font=("Caslon",12),background="#ffffe4")
		txt2 = tk.Text(self,font=("Caslon",16),width=48,height=2) #roll num
		label4=tk.Label(self,text="Enter the ID Card Number of the Student to delete his record! ",font=("Caslon",12),background="#ffffe4")
		txt3 = tk.Text(self,font=("Caslon",16),width=48,height=2) #id num
		bt2=tk.Button(self,text="Continue ",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",12),height=1,width=25,default='active',command=lambda:self.delrecord(txt1.get("1.0",tk.END),txt2.get("1.0",tk.END),txt3.get("1.0",tk.END),controller))
		bt1=tk.Button(self,text="Cancel and go back",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",12),height=1,width=25,default='active',command=lambda:controller.show_frame(StartPage))
		label1.pack(padx=10,pady=25)
		label2.pack(padx=10,pady=25)
		txt1.pack(padx=10,pady=25)
		label3.pack(padx=10,pady=25)
		txt2.pack(padx=10,pady=25)
		label4.pack(padx=10,pady=25)
		txt3.pack(padx=10,pady=25)
		bt2.pack(padx=10,pady=25)
		bt1.pack(padx=10,pady=25)
		
	def delrecord(self,t,u,v,controller):


		path0 = r"C:\Users\kartik\Desktop\GUI stuff\GUIproj.xlsx"
		wb_obj0 = openpyxl.load_workbook(path0) 
		sheet = wb_obj0.active 
		

		for i in range(2,200):
			cell_obj_name = sheet.cell(row = i, column = 1 ).value
			cell_obj_roll = sheet.cell(row = i, column = 4).value
			cell_obj_idnum = sheet.cell(row = i, column = 3).value
			
			if(cell_obj_name == t) and (cell_obj_roll == u) and (cell_obj_idnum == v):
				sheet.cell(row = i, column = 4).value = None
				sheet.cell(row = i, column = 1 ).value = None
				sheet.cell(row = i, column = 3).value = None
				sheet.cell(row = i, column = 2).value = None
				
				break
			else:
				continue				
		wb_obj0.save(path0)

		path1 = r"C:\Users\kartik\Desktop\GUI stuff\SNS.xlsx"
		path2 = r"C:\Users\kartik\Desktop\GUI stuff\OOPS.xlsx"
		path3 = r"C:\Users\kartik\Desktop\GUI stuff\EDC.xlsx"
		path4 = r"C:\Users\kartik\Desktop\GUI stuff\LNT.xlsx"
		path5 = r"C:\Users\kartik\Desktop\GUI stuff\MATHS.xlsx"
		wb_obj1 = openpyxl.load_workbook(path1) 
		wb_obj2 = openpyxl.load_workbook(path2) 
		wb_obj3 = openpyxl.load_workbook(path3) 
		wb_obj4 = openpyxl.load_workbook(path4) 
		wb_obj5 = openpyxl.load_workbook(path5) 
		sheet1 = wb_obj1.active
		sheet2 = wb_obj2.active
		sheet3 = wb_obj3.active
		sheet4 = wb_obj4.active
		sheet5 = wb_obj5.active

		for i in range(2,200):
			cell_obj_name = sheet1.cell(row = i, column = 1 ).value
			cell_obj_roll = sheet1.cell(row = i, column = 4).value

			cell_obj_idnum = (sheet1.cell(row = i, column = 3).value)
			
			if(cell_obj_name == t) and (cell_obj_roll == u) and (cell_obj_idnum == v):
				sheet1.cell(row = i, column = 4).value = None
				sheet1.cell(row = i, column = 1 ).value = None
				sheet1.cell(row = i, column = 3).value = None
				sheet1.cell(row = i, column = 2).value = None
				break
			else:
				continue
		wb_obj1.save(path1)
		for i in range(2,200):
			cell_obj_name = sheet2.cell(row = i, column = 1 ).value
			cell_obj_roll = sheet2.cell(row = i, column = 4).value
			cell_obj_idnum = (sheet2.cell(row = i, column = 3).value)
			
			if(cell_obj_name == t) and (cell_obj_roll == u) and (cell_obj_idnum == v):
				sheet2.cell(row = i, column = 4).value = None
				sheet2.cell(row = i, column = 1 ).value = None
				sheet2.cell(row = i, column = 3).value = None
				sheet2.cell(row = i, column = 2).value = None
				break
			else:
				continue
		wb_obj1.save(path2)
		for i in range(2,200):
			cell_obj_name = sheet3.cell(row = i, column = 1 ).value
			cell_obj_roll = sheet3.cell(row = i, column = 4).value

			cell_obj_idnum = (sheet3.cell(row = i, column = 3).value)
			
			if(cell_obj_name == t) and (cell_obj_roll == u) and (cell_obj_idnum == v):
				sheet3.cell(row = i, column = 4).value = None
				sheet3.cell(row = i, column = 1 ).value = None
				sheet3.cell(row = i, column = 3).value = None
				sheet3.cell(row = i, column = 2).value = None
				break
			else:
				continue
		wb_obj1.save(path3)
		for i in range(2,200):
			cell_obj_name = sheet4.cell(row = i, column = 1 ).value
			cell_obj_roll = sheet4.cell(row = i, column = 4).value

			cell_obj_idnum = (sheet.cell(row = i, column = 3).value)
			
			if(cell_obj_name == t) and (cell_obj_roll == u) and (cell_obj_idnum == v):
				sheet4.cell(row = i, column = 4).value = None
				sheet4.cell(row = i, column = 1 ).value = None
				sheet4.cell(row = i, column = 3).value = None
				sheet4.cell(row = i, column = 2).value = None
				break
			else:
				continue
		wb_obj1.save(path4)
		for i in range(2,200):
			cell_obj_name = sheet5.cell(row = i, column = 1 ).value
			cell_obj_roll = sheet5.cell(row = i, column = 4).value

			cell_obj_idnum = (sheet5.cell(row = i, column = 3).value)
			
			if(cell_obj_name == t) and (cell_obj_roll == u) and (cell_obj_idnum == v):
				sheet5.cell(row = i, column = 4).value = None
				sheet5.cell(row = i, column = 1 ).value = None
				sheet5.cell(row = i, column = 3).value = None
				sheet5.cell(row = i, column = 2).value = None

				controller.show_frame(DeleteRecord)
				messagebox.showinfo("Successful!","Student Record deleted successfully ")
				break
			else:
				continue
		
		wb_obj1.save(path5)	
		
class AddSubjects(tk.Frame):
	def __init__(self,parent,controller):
		tk.Frame.__init__(self,parent)
		self.configure(background="#ffffe4")
		label1=tk.Label(self,text="Add Student's Name",font=("Caslon",16),background="#ffffe4")
		txt1=tk.Text(self,font=("Caslon",16),width=30,height=2)
		label5=tk.Label(self,text="Enter Date of joining",font=("Caslon",16),background="#ffffe4")
		txt5=tk.Text(self,font=("Caslon",16),width=30,height=2)
		label2 =tk.Label(self,text = "Enter Your Roll Number",font = ("Caslon",16),background="#ffffe4")
		txt2 = tk.Text(self,font=("Caslon",16),width = 30,height = 2)
		label3 =tk.Label(self,text = "Enter Id Card number",font = ("Caslon",16),background="#ffffe4")
		txt3 = tk.Text(self,font=("Caslon",16),width = 30,height = 2)
		label4 =tk.Label(self,text = "Enter Surname",font = ("Caslon",16),background="#ffffe4")
		txt4 = tk.Text(self,font=("Caslon",16),width = 30,height = 2)
		bt2=tk.Button(self,text="Add Entry",bg="orange",font=("Caslon",12),height=1,width=17,default='active',command=lambda:[self.addsub(txt1.get("1.0",tk.END),txt4.get("1.0",tk.END),txt3.get("1.0",tk.END),txt2.get("1.0",tk.END),txt5.get("1.0",tk.END),controller),self.speaker('Data added successfully!')])
		bt1=tk.Button(self,text="Back to home",bg="#2983cc",fg='white', cursor='boat',font=("Caslon",12),height=1,width=17,default='active',command=lambda:controller.show_frame(StartPage))
		label1.pack(padx=10,pady=8)
		txt1.pack(padx=10,pady=8)
		label4.pack(padx=10,pady=8)
		txt4.pack(padx=10,pady=8)
		label2.pack(padx=10,pady=8)
		txt2.pack(padx=10,pady=8)
		label3.pack(padx=10,pady=8)
		txt3.pack(padx=10,pady=8)
		label5.pack(padx=10,pady=8)
		txt5.pack(padx=10,pady=8)
		bt2.pack(padx=10,pady=8)	
		bt1.pack(padx=10,pady=8)
	def speaker(self,text):
		pyttsx3.speak(text)
			

	def addsub(self,a,b,c,d,e,controller):
		
		path0 = r"C:\Users\kartik\Desktop\GUI stuff\GUIproj.xlsx"
		path1 = r"C:\Users\kartik\Desktop\GUI stuff\SNS.xlsx"
		path2 = r"C:\Users\kartik\Desktop\GUI stuff\OOPS.xlsx"
		path3 = r"C:\Users\kartik\Desktop\GUI stuff\EDC.xlsx"
		path4 = r"C:\Users\kartik\Desktop\GUI stuff\LNT.xlsx"
		path5 = r"C:\Users\kartik\Desktop\GUI stuff\MATHS.xlsx"

		wb_obj0 = openpyxl.load_workbook(path0)
		wb_obj1 = openpyxl.load_workbook(path1) 
		wb_obj2 = openpyxl.load_workbook(path2) 
		wb_obj3 = openpyxl.load_workbook(path3) 
		wb_obj4 = openpyxl.load_workbook(path4) 
		wb_obj5 = openpyxl.load_workbook(path5) 
		sheet = wb_obj0.active
		sheet1 = wb_obj1.active 
		sheet2 = wb_obj2.active 
		sheet3 = wb_obj3.active 
		sheet4 = wb_obj4.active  
		sheet5 = wb_obj5.active 
		st1 = 'A'
		st2 = 'B'
		st3 = 'C'
		st4 = 'D'
		st5 = 'E'
		ed0 = str(sheet.max_row + 1)
		if (len(a)==1 and a[0]=="") or ((len(b)==1 and b[0]=="")) or ((len(c)==1 and c[0]=="")) or ((len(d)==1 and d[0]=="")) or ((len(e)==1 and e[0]=="")):
			messagebox.showinfo("Alert!", "Please enter the Students")
		
		else:
			cell1 = sheet[st1 + ed0]
			cell1.value = a
			cell2 = sheet[st2 + ed0]
			cell2.value = b
			cell3 = sheet[st3 + ed0]
			cell3.value  = c
			cell4 = sheet[st4 + ed0]
			cell4.value  = d
			cell5 = sheet[st5 + ed0]
			cell5.value  = e
			wb_obj0.save(path0)
			
		ed1 = str(sheet1.max_row + 1)
		if (len(a)==1 and a[0]=="") or ((len(b)==1 and b[0]=="")) or ((len(c)==1 and c[0]=="")) or ((len(d)==1 and d[0]=="")) or ((len(e)==1 and e[0]=="")):
			messagebox.showinfo("Alert!", "Please enter the Students")
		
		else:
			cell1 = sheet1[st1 + ed1]
			cell1.value = a
			cell2 = sheet1[st2 + ed1]
			cell2.value = b
			cell3 = sheet1[st3 + ed1]
			cell3.value  = c
			cell4 = sheet1[st4 + ed1]
			cell4.value  = d
			cell5 = sheet1[st5 + ed1]
			cell5.value  = e
			wb_obj1.save(path1)

		ed2 = str(sheet2.max_row + 1)
		if (len(a)==1 and a[0]=="") or ((len(b)==1 and b[0]=="")) or ((len(c)==1 and c[0]=="")) or ((len(d)==1 and d[0]=="")) or ((len(e)==1 and e[0]=="")):
			messagebox.showinfo("Alert!", "Please enter the Students")
		
		else:
			cell1 = sheet2[st1 + ed2]
			cell1.value = a
			cell2 = sheet2[st2 + ed2]
			cell2.value = b
			cell3 = sheet2[st3 + ed2]
			cell3.value  = c
			cell4 = sheet2[st4 + ed2]
			cell4.value  = d
			cell5 = sheet2[st5 + ed2]
			cell5.value  = e
			wb_obj2.save(path2)
		
		ed3 = str(sheet3.max_row + 1)
		if (len(a)==1 and a[0]=="") or ((len(b)==1 and b[0]=="")) or ((len(c)==1 and c[0]=="")) or ((len(d)==1 and d[0]=="")) or ((len(e)==1 and e[0]=="")):
			messagebox.showinfo("Alert!", "Please enter the Students")
		
		else:
			cell1 = sheet3[st1 + ed3]
			cell1.value = a
			cell2 = sheet3[st2 + ed3]
			cell2.value = b
			cell3 = sheet3[st3 + ed3]
			cell3.value  = c
			cell4 = sheet3[st4 + ed3]
			cell4.value  = d
			cell5 = sheet3[st5 + ed3]
			cell5.value  = e
			wb_obj3.save(path3)
		
		ed4 = str(sheet4.max_row + 1)
		if (len(a)==1 and a[0]=="") or ((len(b)==1 and b[0]=="")) or ((len(c)==1 and c[0]=="")) or ((len(d)==1 and d[0]=="")) or ((len(e)==1 and e[0]=="")):
			messagebox.showinfo("Alert!", "Please enter the Students")
		
		else:
			cell1 = sheet4[st1 + ed4]
			cell1.value = a
			cell2 = sheet4[st2 + ed4]
			cell2.value = b
			cell3 = sheet4[st3 + ed4]
			cell3.value  = c
			cell4 = sheet4[st4 + ed4]
			cell4.value  = d
			cell5 = sheet4[st5 + ed4]
			cell5.value  = e
			wb_obj4.save(path4)
		
		ed5 = str(sheet5.max_row + 1)
		if (len(a)==1 and a[0]=="") or ((len(b)==1 and b[0]=="")) or ((len(c)==1 and c[0]=="")) or ((len(d)==1 and d[0]=="")) or ((len(e)==1 and e[0]=="")):
			messagebox.showinfo("Alert!", "Please enter the Students")
		
		else:
			cell1 = sheet5[st1 + ed5]
			cell1.value = a
			cell2 = sheet5[st2 + ed5]
			cell2.value = b
			cell3 = sheet5[st3 + ed5]
			cell3.value  = c
			cell4 = sheet5[st4 + ed5]
			cell4.value  = d
			cell5 = sheet5[st5 + ed5]
			cell5.value  = e
			wb_obj5.save(path5)
			messagebox.showinfo("Successfully Added!","The Entries are added Successfully.")
			
def main():
	app=AttendanceManager()
	app.title("Attendance System using RFID")	
	app.mainloop()


if __name__=="__main__":
	main()


def validateLogin(username, password):
	print("username entered :", username.get())
	print("password entered :", password.get())
	return


